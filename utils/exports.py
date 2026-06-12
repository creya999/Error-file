import io
from datetime import datetime
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                Paragraph, Spacer)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ─── Brand colours ────────────────────────────────────────────────────────────
ADBL_GREEN  = colors.HexColor('#1a6b3c')
ADBL_LIGHT  = colors.HexColor('#e8f5e9')
ADBL_WHITE  = colors.white
ADBL_DARK   = colors.HexColor('#1a1a1a')


def export_requests_pdf(requests_list, status_label='Approved'):
    """Return a BytesIO PDF of the card request list."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=15*mm, bottomMargin=15*mm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('title', fontSize=14, textColor=ADBL_GREEN,
                                 alignment=TA_CENTER, fontName='Helvetica-Bold')
    sub_style   = ParagraphStyle('sub', fontSize=9, textColor=ADBL_DARK,
                                 alignment=TA_CENTER)

    elements = []

    # Header
    elements.append(Paragraph('Agricultural Development Bank Limited (ADBL)', title_style))
    elements.append(Spacer(1, 3*mm))
    elements.append(Paragraph(
    '<b>Head Office – Digital Banking Department</b>',
    ParagraphStyle(
        'DeptStyle',
        parent=styles['Normal'],
        fontSize=12,
        leading=16,
        alignment=1
    )
))

    elements.append(Paragraph(
    '<b>Instant Card Request List</b>',
    ParagraphStyle(
        'DeptStyle',
        parent=styles['Normal'],
        fontSize=12,
        leading=16,
        alignment=1
    )
))
 
 

    elements.append(Paragraph(f'<b>Generated on: {datetime.now().strftime("%Y-%m-%d %H:%M")}</b>', sub_style))
    elements.append(Spacer(1, 8*mm))

    # Table headers
    headers = ['S.No', 'Request No', 'Branch Code', 'Branch Name',
               'Staff ID', 'Staff Phone', 'Qty', 'Remarks', 'Requested Date', 'Status']

    data = [headers]
    for i, r in enumerate(requests_list, 1):
        data.append([
            str(i),
            r.request_no,
            r.branch_code,
            r.branch_name,
            r.staff_id,
            r.staff_phone,
            str(r.quantity),
            r.remarks or '',
            r.requested_at.strftime('%Y-%m-%d %H:%M') if r.requested_at else '',
            r.status,
        ])

    col_widths = [10*mm, 35*mm, 22*mm, 55*mm, 22*mm, 28*mm, 12*mm, 50*mm, 35*mm, 22*mm]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0),  ADBL_GREEN),
        ('TEXTCOLOR',     (0, 0), (-1, 0),  ADBL_WHITE),
        ('FONTNAME',      (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1, 0),  8),
        ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE',      (0, 1), (-1, -1), 7),
        ('ROWBACKGROUNDS',(0, 1), (-1, -1), [ADBL_WHITE, ADBL_LIGHT]),
        ('GRID',          (0, 0), (-1, -1), 0.3, colors.HexColor('#cccccc')),
        ('LEFTPADDING',   (0, 0), (-1, -1), 3),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 3),
        ('TOPPADDING',    (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    elements.append(table)

    # Footer summary
    total_qty = sum(r.quantity for r in requests_list)
    elements.append(Spacer(1, 4*mm))
    elements.append(Paragraph(
        f'Total Requests: {len(requests_list)}  |  Total Cards Requested: {total_qty}',
        ParagraphStyle('footer', fontSize=9, fontName='Helvetica-Bold',
                       textColor=ADBL_GREEN, alignment=TA_LEFT)
    ))

    doc.build(elements)
    buf.seek(0)
    return buf


def export_requests_excel(requests_list, status_label='Approved'):
    """Return a BytesIO Excel file of the card request list."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Card Requests - {status_label}'

    # Styles
    green_fill  = PatternFill('solid', fgColor='1a6b3c')
    light_fill  = PatternFill('solid', fgColor='e8f5e9')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    data_font   = Font(size=10)
    center      = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin        = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Bank header rows
    ws.merge_cells('A1:J1')
    ws['A1'] = 'Agricultural Development Bank Limited (ADBL)'
    ws['A1'].font = Font(bold=True, size=16, color='1a6b3c')
    ws['A1'].alignment = center

    ws.merge_cells('A2:J2')
    ws['A2'] = f'Instant Card Request List [{status_label}]   |   Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    ws['A2'].font = Font(size=10)
    ws['A2'].alignment = center

    ws.append([])  # blank row

    # Column headers
    headers = ['S.No', 'Request No', 'Branch Code', 'Branch Name',
               'Staff ID', 'Staff Phone', 'Quantity', 'Remarks', 'Requested Date', 'Status']
    ws.append(headers)
    for col, cell in enumerate(ws[4], 1):
        cell.fill   = green_fill
        cell.font   = header_font
        cell.alignment = center
        cell.border = thin

    # Data rows
    for i, r in enumerate(requests_list, 1):
        row = [
            i,
            r.request_no,
            r.branch_code,
            r.branch_name,
            r.staff_id,
            r.staff_phone,
            r.quantity,
            r.remarks or '',
            r.requested_at.strftime('%Y-%m-%d %H:%M') if r.requested_at else '',
            r.status,
        ]
        ws.append(row)
        fill = light_fill if i % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
        for cell in ws[ws.max_row]:
            cell.fill      = fill
            cell.font      = data_font
            cell.alignment = center
            cell.border    = thin

    # Column widths
    col_widths = [5, 20, 14, 35, 15, 16, 10, 30, 20, 12]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # Summary row
    ws.append([])
    summary_row = ws.max_row + 1
    ws.cell(summary_row, 1, f'Total Requests: {len(requests_list)}')
    ws.cell(summary_row, 1).font = Font(bold=True, size=10, color='1a6b3c')
    ws.cell(summary_row, 7, f'Total Qty: {sum(r.quantity for r in requests_list)}')
    ws.cell(summary_row, 7).font = Font(bold=True, size=10, color='1a6b3c')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
